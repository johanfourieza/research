#!/usr/bin/env python3
"""
02_dagregister_signal.py  --  THoTF paper, qualitative backbone (the VOC dagregister)

Reads the shared dagregister Excel and produces:
  * per-year counts of entries matching disease / Khoesan / medicine term-nets, 1685-1730,
    plus total entries per year (the denominator for an attention rate) -> CSV for the
    event-study of smallpox-attention around 1713;
  * a dump of every disease-matched entry 1709-1716 (full Dutch text) -> notes/, for LLM
    classification of WHO is affected (settler/enslaved/Company/Khoesan) and WHETHER death,
    medicine, Batavia or quarantine is mentioned;
  * a focused dump of medicine/Batavia co-mentions 1711-1715 -> notes/, to evidence (not
    assume) the null on "medicines imported from Batavia to treat the afflicted".

Run: python 02_dagregister_signal.py
"""
import openpyxl, os, re, csv
from collections import defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.normpath(os.path.join(HERE, "..", "..", "sources", "data"))
XLSX = os.path.join(DATA, "vc_datastel (v3).xlsx")
OUT = os.path.join(HERE, "outputs"); os.makedirs(OUT, exist_ok=True)
NOTES = os.path.normpath(os.path.join(HERE, "..", "notes")); os.makedirs(NOTES, exist_ok=True)

# --- term nets (Dutch, period spelling + OCR variants) ------------------------
SMALLPOX = re.compile(r"kinderpok|kinder[- ]?pok|pokjes|pokken|pocken|pokk?en", re.I)
DISEASE  = re.compile(r"kinderpok|pokjes|pokken|pocken|besmettelijk|besmetting|sterfte|"
                      r"epidemi|sieckt|siekte|ziekte|sieckte|kranckt|sterv", re.I)
DEATH    = re.compile(r"\boverle|gestorv|\bsterv|\bdoot\b|\bdood|begrav|lijk|lyck|overleeden", re.I)
KHOE     = re.compile(r"hottent|sonqua|soncqua|bosjesman|bossiesman|inboorling|\bwilden\b", re.I)
SLAVE    = re.compile(r"slaa?v|lijfeigen|mancipia|caffer", re.I)
MEDICINE = re.compile(r"medicij|medicament|geneesmid|\bgenees|chirurg|doctor|docter|apot|"
                      r"balsem|drogerij|remedie|cureren|geneeskund", re.I)
BATAVIA  = re.compile(r"\bbatavia", re.I)
QUAR     = re.compile(r"quarantain|quarentain|afsonder|gesepareer|geweerd|verbod", re.I)

def strip_tags(s): return re.sub(r"<[^>]*>", " ", s)

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]

per_year = defaultdict(lambda: defaultdict(int))
disease_dump, med_dump = [], []
for row in ws.iter_rows(min_row=2, values_only=True):
    dv, entry = row[1], row[3]
    if not dv: continue
    try: y = int(str(dv)[:4])
    except ValueError: continue
    if not (1685 <= y <= 1730): continue
    per_year[y]["total"] += 1
    if not entry or str(entry).strip().lower() in ("", "geen teks"): continue
    e = strip_tags(str(entry))
    is_dis = bool(DISEASE.search(e))
    if is_dis: per_year[y]["disease"] += 1
    if SMALLPOX.search(e): per_year[y]["smallpox"] += 1
    if KHOE.search(e) and is_dis: per_year[y]["khoe_disease"] += 1
    if SLAVE.search(e) and is_dis: per_year[y]["slave_disease"] += 1
    if MEDICINE.search(e): per_year[y]["medicine"] += 1
    if BATAVIA.search(e) and MEDICINE.search(e): per_year[y]["batavia_medicine"] += 1
    # dumps for classification
    if 1709 <= y <= 1716 and is_dis:
        disease_dump.append((str(dv), e.strip()))
    if 1711 <= y <= 1715 and MEDICINE.search(e) and (DISEASE.search(e) or BATAVIA.search(e)):
        med_dump.append((str(dv), e.strip()))

# --- per-year counts CSV ------------------------------------------------------
cols = ["total", "disease", "smallpox", "khoe_disease", "slave_disease", "medicine", "batavia_medicine"]
with open(os.path.join(OUT, "dagregister_year_counts.csv"), "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f); w.writerow(["year"] + cols + ["disease_rate_per1000"])
    for y in sorted(per_year):
        d = per_year[y]; tot = d["total"] or 1
        w.writerow([y] + [d[c] for c in cols] + [round(1000 * d["disease"] / tot, 2)])

# --- dumps --------------------------------------------------------------------
with open(os.path.join(NOTES, "dagregister_disease_entries_1709_1716.tsv"), "w", encoding="utf-8") as f:
    f.write("date\tentry\n")
    for d, e in disease_dump: f.write(f"{d}\t{e}\n")
with open(os.path.join(NOTES, "dagregister_medicine_comentions_1711_1715.tsv"), "w", encoding="utf-8") as f:
    f.write("date\tentry\n")
    for d, e in med_dump: f.write(f"{d}\t{e}\n")

# --- console report -----------------------------------------------------------
print("year  total  disease  smallpox  khoe+dis  slave+dis  medicine  bat+med")
for y in sorted(per_year):
    d = per_year[y]
    if 1705 <= y <= 1720:
        print(f"{y}  {d['total']:5d}  {d['disease']:7d}  {d['smallpox']:8d}  {d['khoe_disease']:8d}  "
              f"{d['slave_disease']:9d}  {d['medicine']:8d}  {d['batavia_medicine']:7d}")
print(f"\nDisease entries 1709-1716 dumped: {len(disease_dump)}")
print(f"Medicine co-mentions 1711-1715 dumped: {len(med_dump)}")
print("Outputs ->", OUT, "and", NOTES)
