#!/usr/bin/env python3
"""
05_primary_classifier.py  --  the transparent "cheap" primary classifier over the FULL
1700-1720 dagregister population, per the frozen codebook, plus a random validation draw.

Implements V1 (disease/death), V3 (smallpox), V4 (epidemic medicine/Batavia) and a coarse V2
(who) by keyword rules. These are the LLM/primary labels Vhat_r to be DEBIASED against a
hand/gold-labelled random validation sample (w33344 §4.3). Deterministic & reproducible.

Outputs:
  outputs/primary_labels_1700_1720.csv      -- every entry with id,date,year,has_text,V1,V3,V4,who,evidence
  notes/validation_sample_1700_1720.tsv     -- random N=400 entries (id,date,entry) for GOLD labelling
  outputs/primary_year_rates.csv            -- plug-in yearly rates of V1,V3,V4
Run: python 05_primary_classifier.py
"""
import openpyxl, os, re, csv, hashlib

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.normpath(os.path.join(HERE, "..", "..", "sources", "data"))
XLSX = os.path.join(DATA, "vc_datastel (v3).xlsx")
OUT = os.path.join(HERE, "outputs"); os.makedirs(OUT, exist_ok=True)
NOTES = os.path.normpath(os.path.join(HERE, "..", "notes")); os.makedirs(NOTES, exist_ok=True)
N_VALID = 400

# --- term nets (frozen) -------------------------------------------------------
SMALLPOX = re.compile(r"kinderpok|kinder[- ]?pok|pokjes|pokken|pocken", re.I)
DEATH    = re.compile(r"\boverle|gestorv|\bsterv|sterft|\bdoot\b|\bdood|begrav|lijk\b|lijkk|lyck", re.I)
SICK     = re.compile(r"sieck|siek|ziek|kranck|krank|koorts|besmettelijk|besmetting|epidemi|pest\b", re.I)
LIVESTOCK= re.compile(r"\bvee\b|beesten|rundvee|schapen|bestiaal|trekbeest", re.I)
KHOE     = re.compile(r"hottent|sonqua|soncqua|bosjesman|bossiesman|inboorling|\bwilden\b|afrikaan", re.I)
CSLAVE   = re.compile(r"comp\w*\W+(s\W+)?(slav|lijfeig|mancip)|'s\s*comp", re.I)
SLAVE    = re.compile(r"slaa?v|lijfeig|mancipia", re.I)
SOLDSAIL = re.compile(r"matroos|matrosen|soldaat|soldaten|bootsgesel|scheeps|koppen|afgevare|equipage", re.I)
SETTLER  = re.compile(r"burger|coloni|landbouw|ingeseten|ingelande|vrijman|vrylieden|inwoonder", re.I)
MED_EPI  = re.compile(r"medicij|medicament|geneesmid|geneesing|genees\w*|chirurg|doctor|docter|apot|"
                      r"balsem|drogerij|remedie|cureer|quarantain|quarentain", re.I)
BATAVIA  = re.compile(r"\bbatavia", re.I)

def strip(s): return re.sub(r"<[^>]*>", " ", str(s))

def span(rx, t, w=60):
    m = rx.search(t)
    if not m: return ""
    a = max(0, m.start()-w); b = min(len(t), m.end()+w)
    return re.sub(r"\s+", " ", t[a:b]).strip()

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    rid, dv, day, entry = r[0], r[1], r[2], r[3]
    if not dv: continue
    try: y = int(str(dv)[:4])
    except ValueError: continue
    if not (1700 <= y <= 1720): continue
    has_text = bool(entry) and str(entry).strip().lower() not in ("", "geen teks")
    t = strip(entry) if has_text else ""
    # V1: human disease/death (exclude pure-livestock death mentions)
    human_death = bool(DEATH.search(t)) and not (LIVESTOCK.search(t) and not (SLAVE.search(t) or KHOE.search(t) or SOLDSAIL.search(t) or SETTLER.search(t)))
    v1 = 1 if (SICK.search(t) or SMALLPOX.search(t) or human_death) else 0
    v3 = 1 if SMALLPOX.search(t) else 0
    v4 = 1 if (v1 and (MED_EPI.search(t) or (BATAVIA.search(t) and (SICK.search(t) or SMALLPOX.search(t))))) else 0
    who = []
    if v1:
        if CSLAVE.search(t): who.append("company_slaves")
        elif SLAVE.search(t): who.append("private_slaves")
        if KHOE.search(t): who.append("khoesan")
        if SOLDSAIL.search(t): who.append("soldiers_sailors")
        if SETTLER.search(t): who.append("settlers")
        if not who: who.append("general")
    rows.append({"id": rid, "date": str(dv), "year": y, "has_text": int(has_text),
                 "V1": v1, "V3": v3, "V4": v4, "who": "|".join(who),
                 "evidence": span(SMALLPOX if v3 else (DEATH if human_death else SICK), t)})

# write per-entry labels
with open(os.path.join(OUT, "primary_labels_1700_1720.csv"), "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=["id","date","year","has_text","V1","V3","V4","who","evidence"])
    w.writeheader(); w.writerows(rows)

# plug-in yearly rates
from collections import defaultdict
yr = defaultdict(lambda: [0,0,0,0])  # n, V1, V3, V4
for d in rows:
    a = yr[d["year"]]; a[0]+=1; a[1]+=d["V1"]; a[2]+=d["V3"]; a[3]+=d["V4"]
with open(os.path.join(OUT, "primary_year_rates.csv"), "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f); w.writerow(["year","n","V1_disease","V3_smallpox","V4_medicine","V1_rate","V3_rate"])
    for y in sorted(yr):
        n,v1,v3,v4 = yr[y]
        w.writerow([y,n,v1,v3,v4, round(v1/n,4), round(v3/n,4)])

# deterministic random validation sample (seeded by id hash, no Date/random needed)
def h(s): return int(hashlib.md5(str(s).encode()).hexdigest(), 16)
ranked = sorted(rows, key=lambda d: h(d["id"]))
valid = ranked[:N_VALID]
with open(os.path.join(NOTES, "validation_sample_1700_1720.tsv"), "w", encoding="utf-8") as f:
    f.write("id\tdate\tyear\tprimary_V1\tprimary_V3\tprimary_V4\tprimary_who\tentry\n")
    # re-read full text for the sampled ids
    sampled_ids = {d["id"] for d in valid}
    textmap = {}
    ws2 = wb[wb.sheetnames[0]]
    for r in ws2.iter_rows(min_row=2, values_only=True):
        if r[0] in sampled_ids:
            textmap[r[0]] = strip(r[3]) if (r[3] and str(r[3]).strip().lower() not in ("","geen teks")) else "(no entry)"
    for d in valid:
        f.write(f"{d['id']}\t{d['date']}\t{d['year']}\t{d['V1']}\t{d['V3']}\t{d['V4']}\t{d['who']}\t{textmap.get(d['id'],'')}\n")

# console
print(f"Entries 1700-1720: {len(rows)}  (text: {sum(d['has_text'] for d in rows)})")
print("Plug-in yearly rates (year: n V1 V3 V4):")
for y in sorted(yr):
    n,v1,v3,v4 = yr[y]; print(f"  {y}: n={n:3d}  V1={v1:3d}  V3(smallpox)={v3:3d}  V4(medicine)={v4:2d}")
print(f"\nValidation sample (N={N_VALID}) -> notes/validation_sample_1700_1720.tsv  (for gold labelling)")
print("Per-entry primary labels -> outputs/primary_labels_1700_1720.csv")
