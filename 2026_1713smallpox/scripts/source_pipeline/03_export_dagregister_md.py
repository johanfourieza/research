#!/usr/bin/env python3
"""
03_export_dagregister_md.py  --  export the dagregister window to a readable markdown corpus,
one file per year (1700-1720), each day a dated section. This lets agents read whole years in
full (no regex snippets), so the qualitative analysis can be as complete as possible.

Output: THoTF/notes/dagregister_md/<year>.md  + index.md
Run: python 03_export_dagregister_md.py
"""
import openpyxl, os, re
from collections import defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.normpath(os.path.join(HERE, "..", "..", "sources", "data"))
XLSX = os.path.join(DATA, "vc_datastel (v3).xlsx")
OUT = os.path.normpath(os.path.join(HERE, "..", "notes", "dagregister_md")); os.makedirs(OUT, exist_ok=True)

def clean(s):
    s = re.sub(r"<br\s*/?>", "\n", str(s))
    s = re.sub(r"<p\s*/?>", "\n\n", s)
    s = re.sub(r"<[^>]*>", " ", s)
    s = s.replace("�", "'")  # OCR replacement char -> apostrophe (period spelling)
    return re.sub(r"[ \t]+", " ", s).strip()

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
by_year = defaultdict(list)
for row in ws.iter_rows(min_row=2, values_only=True):
    rid, dv, day, entry = row[0], row[1], row[2], row[3]
    if not dv: continue
    try: y = int(str(dv)[:4])
    except ValueError: continue
    if 1700 <= y <= 1720:
        by_year[y].append((str(dv), str(day) if day else "", entry))

counts = {}
for y in sorted(by_year):
    lines = [f"# Dagregister van de Caab — {y}\n",
             "Daily entries of the VOC Council of Policy / governor's journal (Dutch, period "
             "spelling; OCR-cleaned). Each section is one day.\n"]
    n_text = 0
    for dv, day, entry in by_year[y]:
        lines.append(f"\n## {dv}  {day}\n")
        if not entry or str(entry).strip().lower() in ("", "geen teks"):
            lines.append("_(no entry)_\n")
        else:
            lines.append(clean(entry) + "\n")
            n_text += 1
    with open(os.path.join(OUT, f"{y}.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    counts[y] = (len(by_year[y]), n_text)

with open(os.path.join(OUT, "index.md"), "w", encoding="utf-8") as f:
    f.write("# Dagregister markdown corpus (1700-1720)\n\n")
    f.write("One file per year; read whole years to extract disease/death/Khoesan/slave/"
            "medicine/Batavia/quarantine evidence in full.\n\n| Year | days | days_with_text |\n|---|---|---|\n")
    for y in sorted(counts):
        f.write(f"| [{y}]({y}.md) | {counts[y][0]} | {counts[y][1]} |\n")

print("Wrote per-year markdown to", OUT)
for y in sorted(counts): print(f"  {y}.md  days={counts[y][0]}  with_text={counts[y][1]}")
