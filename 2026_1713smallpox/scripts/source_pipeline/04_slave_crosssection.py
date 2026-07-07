#!/usr/bin/env python3
"""
04_slave_crosssection.py  --  RAW cross-sectional enslaved counts by district, year and age/sex
category, straight from the opgaafrolle year-sheets (NO panel, NO matching).

Per the standard opgaaf layout, each year-sheet has slave columns:
  Slaven (adult men) | Slavinnen (adult women) | Jongens (boys) | Meijsies (girls)
plus settler columns Mannen/Vrouwen/Zoonen/Dogters and Knegts. We sum over genuine data rows
(those whose 'Nr.' is an integer), skipping folio/subtotal rows. If the enslaved totals fall
in 1713/1714, the epidemic touched the enslaved; if not, that is itself a striking finding.

Output: outputs/slave_crosssection_by_district_year.csv  (+ console table)
Run: python 04_slave_crosssection.py
"""
import openpyxl, os, csv

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.normpath(os.path.join(HERE, "..", "..", "sources", "data"))
OUT = os.path.join(HERE, "outputs"); os.makedirs(OUT, exist_ok=True)
FILES = {
    "Cape": "Early Cape District - Including Full Index for Hague & Cape Archives September 2024.xlsx",
    "StelDrak": "Stellenbosch-Drakenstein Earlier Opgaafrolle - Hague & Cape Archives - including Full Indexes June 2022.xlsx",
}
# logical column -> set of header-label variants seen across year-sheets
SYN = {
    "slave_men":   {"slaven"},
    "slave_women": {"slavinnen"},
    "slave_boys":  {"jongens", "jongetjes"},
    "slave_girls": {"meijsies", "meijsjes", "meisjes"},
    "set_men":     {"mannen", "mans"},
    "set_women":   {"vrouwen"},
    "set_sons":    {"zoonen", "soons", "soonen"},
    "set_daught":  {"dogters", "dochters"},
    "knechts":     {"knegts", "knechts"},
    "nr":          {"nr.", "nr"},
}
SLAVE = ["slave_men", "slave_women", "slave_boys", "slave_girls"]
SETTLER = ["set_men", "set_women", "set_sons", "set_daught"]

def num(x):
    try: return float(x)
    except (TypeError, ValueError): return 0.0

def is_int(x):
    try: int(str(x).strip()); return True
    except (TypeError, ValueError): return False

def find_header(rows):
    """Return (row_idx, {logical_name: col_idx}) by matching header-label synonyms."""
    for i, r in enumerate(rows[:8]):
        labels = [str(c).strip().lower() if c is not None else "" for c in r]
        if "slavinnen" in labels and ("mannen" in labels or "mans" in labels):
            cmap = {}
            for logical, variants in SYN.items():
                for j, lab in enumerate(labels):
                    if lab in variants:
                        cmap[logical] = j; break
            return i, cmap
    return None, None

records = []
for dist, fn in FILES.items():
    wb = openpyxl.load_workbook(os.path.join(DATA, fn), read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        if not (sheet.isdigit() and 1700 <= int(sheet) <= 1720): continue
        ws = wb[sheet]
        allrows = list(ws.iter_rows(values_only=True))
        hi, colmap = find_header(allrows)
        if colmap is None: continue
        agg = {c: 0.0 for c in SLAVE + SETTLER + ["knechts"]}
        ndata = 0
        nr_col = colmap.get("nr", 0)
        for r in allrows[hi+1:]:
            if nr_col >= len(r) or not is_int(r[nr_col]): continue
            ndata += 1
            for c in agg:
                j = colmap.get(c)
                if j is not None and j < len(r): agg[c] += num(r[j])
        sl_men = agg["slave_men"]; sl_wom = agg["slave_women"]
        sl_boy = agg["slave_boys"]; sl_girl = agg["slave_girls"]
        sl_tot = sl_men + sl_wom + sl_boy + sl_girl
        set_tot = sum(agg[c] for c in SETTLER)
        records.append([dist, int(sheet), ndata, int(sl_men), int(sl_wom), int(sl_boy),
                        int(sl_girl), int(sl_tot), int(agg["knechts"]), int(set_tot),
                        int(agg["set_men"]), int(agg["set_women"]),
                        int(agg["set_sons"]), int(agg["set_daught"])])

records.sort(key=lambda x: (x[0], x[1]))
hdr = ["district", "year", "n_households", "slave_men", "slave_women", "slave_boys",
       "slave_girls", "slave_total", "knechts", "settler_total",
       "settler_men", "settler_women", "settler_sons", "settler_daught"]
with open(os.path.join(OUT, "slave_crosssection_by_district_year.csv"), "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f); w.writerow(hdr); w.writerows(records)

# console: per district, the trajectory + YoY % change in slave_total
for dist in FILES:
    rs = [r for r in records if r[0] == dist]
    print(f"\n=== {dist} district — raw opgaaf cross-sections ===")
    print("year  nHH  slaveM slaveW slaveB slaveG  slaveTOT  YoY%   settlerTOT  set_YoY%")
    prev = prevset = None
    for r in rs:
        st = r[7]; se = r[9]
        yo = f"{(st/prev-1)*100:+6.1f}" if prev else "    . "
        so = f"{(se/prevset-1)*100:+6.1f}" if prevset else "    . "
        print(f"{r[1]}  {r[2]:4d}  {r[3]:6d} {r[4]:6d} {r[5]:6d} {r[6]:6d}  {st:8d}  {yo}   {se:8d}  {so}")
        prev, prevset = st, se
print("\nWritten ->", os.path.join(OUT, "slave_crosssection_by_district_year.csv"))
